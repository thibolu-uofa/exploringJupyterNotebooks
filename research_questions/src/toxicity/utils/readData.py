"""
File Summary:

Module is used to gather column, label, or key data of the specified file type. Each function turns the specified
selection within the file into a list of specified data.

"""

#Imports

#Local
from utils import utils

# Built-in
import json
from itertools import islice

# Installed
import openpyxl
import pandas as pd


def getExcelLabels(fileName : str, sheetName : str = None, specifyFilePath : bool = False) -> list:

    # Check if user wants to specify file path
    if specifyFilePath:
        # Read excel workbook
        wb = openpyxl.load_workbook(fileName)
    else:
        # Read excel workbook
        wb = openpyxl.load_workbook(f"data/results/{fileName}.xlsx")

    # Check if sheetName is specified
    if sheetName:
        # Get the specified sheet
        rawSheet = wb[sheetName]
    else:
        # Get the first sheet
        rawSheet = wb.active

    for row in rawSheet.iter_rows(max_row=1, min_col=1):
        return [cell.value for cell in row]


def getExcelData(fileName : str, *column : str, sheetName : str = None, ignoreLabels : bool = True ,
                 specifyFilePath : bool = False) -> list[list]:
    '''
    Obtains any amount of specified column data in an Excel file. Each column is turned into a list and further returned
    as a group of the specified columns in order.

    :param fileName: Name of the Excel file, without the .xlsx file extension.
    :param column: Accepts any amount of columns to turn into a list. Specified by column letter
    :param sheetName: Optional argument to specify the sheet name of the Excel sheet
    :param ignoreLabels: Check if the user wants to remove the labels of a Excel column or not. Default is True to
                         remove labels
    :param specifyFilePath: Boolean for if the user wants to specify the file path of the read file. Use fileName parameter
                            as the file path location.
    :return: Returns a list of the grouped column data in specified order
    '''

    # Initialize list containing each column data
    tempExcelList : list = []

    # Check if user wants to specify file path
    if specifyFilePath:
        # Read excel workbook
        wb = openpyxl.load_workbook(fileName)
    else:
        # Read excel workbook
        wb = openpyxl.load_workbook(f"data/results/{fileName}.xlsx")

    # Check if sheetName is specified
    if sheetName:
        # Get the specified sheet
        rawSheet = wb[sheetName]
    else :
        # Get the first sheet
        rawSheet = wb.active

    # Check if needed to ignore the first row
    if ignoreLabels:

        # Iterate though each row
        for col in column:

            # Add each column to temp list
            tempExcelList.append([cell.value for cell in islice(rawSheet[col], 1, None)])
    else:

        # Iterate through each row
        for col in column:

            # Add each column to temp list
            tempExcelList.append([cell.value for cell in rawSheet[col]])

    #Return all data as a list
    return tempExcelList

def getCsvData(fileName : str, *labels : str , specifyFilePath : bool = False) -> list[list]:
    '''
    Obtains any amount of specified label data in an Excel file. Each label is turned into a list and further returned
    as a group of the specified label in order.

    :param fileName: Name of the CSV file, without the .csv file extension.
    :param labels: Accepts any amount of labels to turn into a list. Specified by label name
    :param specifyFilePath: Boolean for if the user wants to specify the file path of the read file. Use fileName parameter
                            as the file path location.
    :return: Returns a list of the grouped label data in specified order
    '''

    # Initialize list containing label data
    tempCsvList : list = []

    # Check if the user wants to specify file path
    if specifyFilePath:

        # Read specified CSV file
        df = pd.read_csv(fileName, usecols=[label for label in labels])
    else:

        # Read CSV file
        df = pd.read_csv(f'../data/paperData/{fileName}.csv', usecols=[label for label in labels])

    # Add each label to list
    for label in labels:
        tempCsvList.append(df[label].tolist())

    # Return list of all labels
    return tempCsvList



def getJsonData(fileName : str, *keys : str, specifyFilePath : bool = False) -> list[list]:
    '''
    Obtains any amount of specified key data in an JSON file. Each key is turned into a list and further returned
    as a group of the specified key in order.

    :param fileName: Name of the JSON file, without the .json file extension.
    :param keys:  Accepts any amount of keys to turn into a list. Specified by key name, use '>' to indicate nested
                  dictionary keys
    :param specifyFilePath: Boolean for if the user wants to specify the file path of the read file. Use fileName parameter
                            as the file path location.
    :return: Returns a list of the grouped key data in specified order
    '''

    # Initialize list for each Json category
    tempJsonList : list = []

    # Check if user wants to specify file path
    if specifyFilePath:

        # Open file
        loadFile = open(fileName, "r")
    else:

        # Open file
        loadFile = open(f"./data/results/{fileName}.json", "r")

    # Make a new list for each key
    for _ in keys:
        tempJsonList.append([])

    # Iterate through each line of the JSON file
    for line in set(loadFile):

        # Load JSON data into a dictionary
        convertedData : dict = json.loads(line)

        for keyNumber,key in enumerate(keys):

            # Check if the key is nested
            if ">" in key:
                # Obtain data
                tempJsonList[keyNumber].append(utils.getNestedDictionaryValue(convertedData, key))

                # Skip to next line
                continue

            # Add key data to specified list
            tempJsonList[keyNumber].append(convertedData[key])

    # Close file
    loadFile.close()

    # Return all key data as a list
    return tempJsonList

