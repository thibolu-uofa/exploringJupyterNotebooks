"""
File Summary:

Module is used to create JSON and Excel files based on the provided data.

Creating JSON files are used to mainstream .txt into a more supported file type of JSON.

Excel file creation is used to view data easily.
"""

#Imports

# Built-in
import json

# Installed
import openpyxl
from openpyxl.styles import Alignment


def createJsonFile(saveFileName: str, keys: tuple, data: list[list]) -> None:
    '''
    Used to format txt files into a more supported file type of JSON, so data can be easily identified on each line
    rather than using delimiters.

    Each list size must correspond the same amount of keys.

    :param saveFileName: Save file name to put into the results folder
    :param keys: Each identifying label to organize the JSON object
    :param data: Each list is a JSON object to be put in the JSON file
    :return: Creates a using the saveFileName in the results folder
    '''

    #Open file
    with open(f"../data/evaluations/{saveFileName}.json", "a") as writeFile:

        #Iterate through every list in the data
        for listData in data:

            #Check if there is invalid entries
            if len(listData) != len(keys):
                raise Exception("List size does not equal key size. Each value in each list must correspond to a key!")

            #Intialize the dictionary
            tempDict = dict.fromkeys(keys, None)

            #Iterate through each key and add each value to it
            for keyNumber, key in enumerate(tempDict):
                tempDict[key] = listData[keyNumber]

            #Append to the specified file
            writeFile.write(json.dumps(tempDict) + "\n")



def createExcelFile(sheetTitle: str, label: tuple, data: list[list]):
    """
    Function should be primarily used as a way to view data from a JSON file format

    :param sheetTitle: Title of the worksheet to be applied
    :param label: Form of a tuple, contains labels
    :param data: Data to append to the Excel file in terms of rows
    :return: Creates Excel file
    """

    # WorkBook Save Variables
    workbook = openpyxl.Workbook()

    # Delete default Sheet
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    # Create sheet and use title
    namedSheet = workbook.create_sheet(title=sheetTitle)

    # Add labels
    namedSheet.append(label)

    # Add data to sheet
    for number, row in enumerate(data, start = 2):
        namedSheet.append(row)
        namedSheet['A' + str(number)].alignment = Alignment(wrapText=True)


    # Set Column Width
    for col in namedSheet.columns:
        # Get the column letter
        column = col[0].column_letter

        # Get max length of text
        maxColumnWidth = max([len(str(cell.value).strip()) for cell in namedSheet[column]])

        # Set column width to max text length and limit field to 80
        if maxColumnWidth > 80:
            namedSheet.column_dimensions[column].width = 80
        else:
            namedSheet.column_dimensions[column].width = maxColumnWidth


    # Save file
    workbook.save(f"data/results/{sheetTitle}.xlsx")


def toxicityDataToExcel(saveFileName : str, sentenceData : list[str], evaluatedData : list
                     ,*,extraDataLabels : tuple = None, extraData : list = None, threshold : float = 0.5) -> None:

    '''
    Saves all toxicity data that meets the threshold requirements in an Excel file for easy readability. Keyword
    arguments need to be specified.

    :param saveFileName: Name of the Excel file to save in "./data/results"
    :param extraDataLabels: Corresponding labels to extra data
    :param threshold: Range of 0.00 - 1.00, default is 0.5
    :param extraData: Any data to add in the Excel file, must add a corresponding label to each data except the body
                      label.
    :return: Creates an Excel file containing toxicity data in "./data/results"
    '''

    if len(extraDataLabels) - 1 != len(extraData):
        raise Exception("The label length does not match the provided object data!")


    # Temp list of each excel row data
    tempList : list[list] = []

    # Iterate through each set of evaluated data
    for dataNumber, toxicityData in enumerate(evaluatedData):

        # Check if evaluated data element is valid
        if not toxicityData: continue

        # Check if the evaluated data element meets the threshold requirement
        if toxicityData >= threshold:
            tempList.append([sentenceData[dataNumber], toxicityData])

            # Check if there is extraData to add
            if extraData:
                for data in extraData:
                    tempList[-1].append(data[dataNumber])

    # Save data to Excel file
    createExcelFile(saveFileName, sum((("Sentence",), extraDataLabels), ()), tempList)


